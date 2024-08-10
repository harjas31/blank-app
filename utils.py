import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def export_to_excel(results, file_path, platform):
    workbook = Workbook()
    workbook.remove(workbook.active)  # Remove default sheet

    for keyword, products in results.items():
        sheet = workbook.create_sheet(title=keyword[:31])  # Excel sheet names limited to 31 characters

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        headers = ["Rank", "ASIN" if platform == "Amazon" else "Product ID", "Title", "Price"]
        for col, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment

        # Write data
        for row, product in enumerate(products, start=2):
            sheet.cell(row=row, column=1, value=product["rank"])
            sheet.cell(row=row, column=2, value=product.get("asin") or product.get("product_id"))
            sheet.cell(row=row, column=3, value=product["title"])
            sheet.cell(row=row, column=4, value=product["price"])

    workbook.save(file_path)

def read_keywords_from_excel(file_path):
    try:
        df = pd.read_excel(file_path, header=None)
        keywords = df[0].tolist()
        return [keyword for keyword in keywords if isinstance(keyword, str)]
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []
